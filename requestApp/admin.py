from django.contrib import admin
from django.http import HttpResponse, HttpResponseRedirect
from .models import COLOUser

#ENABLES ABILITY TO EXPORT REQUESTS TO A .CSV DOCUMENT FROM THE ADMIN SITE
def export_csv(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    #set response to download created file
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=LSLDC_COLO_REQUESTS.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))
    writer.writerow([
        smart_str(u"Name"),
        smart_str(u"Email"),
        smart_str(u"Phone Number"),
        smart_str(u"Reason for request"),
        smart_str(u"UCARD ID"),
        smart_str(u"Manager Name"),
        smart_str(u"Manager Email"),
        smart_str(u"Submission Date"),
        smart_str(u"Manager Approval"),
        smart_str(u"COLO Manager Approval"),
    ])
    
    #copy database information to csv
    for obj in queryset:
        writer.writerow([
            smart_str(obj.name),
            smart_str(obj.email),
            smart_str(obj.phone),
			smart_str(obj.reason),
			smart_str(obj.UCard_ID),
			smart_str(obj.manager),
			smart_str(obj.man_email),
			smart_str(obj.time),
			smart_str(obj.man_approved),
			smart_str(obj.COLO_approved),
        ])
    return response
export_csv.short_description = u"Export CSV"

#ENABLES ABILITY TO EXPORT REQUESTS TO AN .XLS DOC FROM THE ADMIN SITE
def export_xls(modeladmin, request, queryset):
    import xlwt
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=LSLSDC_COLO_REQUESTS.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("MyModel")
    
    row_num = 0
    
    columns = [
        (u"Name", 4000),
        (u"Email", 4000),
        (u"Phone", 4000),
    	(u"Reason", 4000),
		(u"UCard_ID", 3000),
		(u"Manager Name", 4000),
		(u"Manager Email", 5000),
		(u"Time Of Submission", 5000),
		(u"Supervisor Approval", 5000),
		(u"COLO Approval", 5000),

    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    
    for obj in queryset:
        row_num += 1
        row = [
            obj.name,
            obj.email,
            obj.phone,
            obj.reason,
            obj.UCard_ID,
            obj.manager,
            obj.man_email,
            obj.man_approved,
            obj.COLO_approved,
        ]
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
            
    wb.save(response)
    return response
export_xls.short_description = u"Export XLS"

#ENABLES ABILITY TO EXPORT REQUESTS TO AN .XLSX DOC FROM THE ADMIN SITE
def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.utils.cell import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=LSLDC_COLO_REQUESTS.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "MyModel"

    row_num = 0

    columns = [
        (u"Name", 30),
        (u"Email", 30),
        (u"Phone", 15),
    	(u"Reason", 30),
		(u"UCard_ID", 15),
		(u"Manager Name", 30),
		(u"Manager Email", 30),
		(u"Time Of Submission", 20),
		(u"Supervisor Approval", 18),
		(u"COLO Approval", 18),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        #c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.name,
            obj.email,
            obj.phone,
            obj.reason,
            obj.UCard_ID,
            obj.manager,
            obj.man_email,
            obj.time,
            obj.man_approved,
            obj.COLO_approved,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]

    wb.save(response)
    return response
export_xlsx.short_description = u"Export XLSX"

#DISABLE EDITING OF REQUESTER INFORMATION
class readAdmin(admin.ModelAdmin):
	#ENABLING THESE LINES DISABLES ABILITY TO CHANGE INFORMATION IN THE ADMIN SITE
	# readonly_fields= (
# 	'name', 
# 	'email', 
# 	'phone', 
# 	'reason',
# 	'UCard_ID',
# 	'manager',
# 	'man_email',
# 	'time', 
# 	'man_approved', 
# 	'COLO_approved',)

	#ENABLE ACTIONS ON THE ADMIN SITE
	actions= [export_csv, export_xls, export_xlsx]

admin.site.register(COLOUser, readAdmin)

